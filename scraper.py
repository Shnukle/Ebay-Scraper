from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook, load_workbook

#requesting info from link and reading it w/ bs4
def scrape_data(url):
    result = requests.get(url)
    doc = BeautifulSoup(result.text, "html.parser")
    return doc
#asks scraping questions    
get_link = input("enter the ebay link of product you want to scrape: ")
base_url = f"{get_link}&_pgn="
num_pages = int(input("How many pages do you want to collect data from?: "))

#creates workbook and A row for descriptions
wb = Workbook()
ws = wb.active
ws.append(["prices", "links", "average price", "maximum price", "minimum price"])


items = []
count = 1
links = []
#scrapes pages requested
for page in range(1, num_pages + 1):
    url = base_url + str(page)
    print(f"page: {count}")
    count += 1
    soup = scrape_data(url)
    products = soup.find("ul", {"class": "srp-results"}).find_all("li", {"class": "s-item"})
    #gets price from every listing on the page
    for product in products:
        cost = product.find("span", {"class": "s-item__price"}).text
        link = product.find("a", {"target": "_blank"}).get("href")
        if 'to' in cost:
            continue
        price = float(cost[1:].replace(",",""))
        #putting data into xlsx file
        ws.append([str(f"${price}"), link])
        items.append(price)
        links.append(link)

#functions to find average price, min price, and max price of products
def get_average():
    average = int(sum(items) / len(items))
    return average

def get_min():
    min = items[0]
    for item in items:
        if item < min:
            min = item
    return min

def get_max():
    max = 0
    for item in items:
        if item > max:
            max = item
    return max

#function to save data into xlsx file 
def save_data_to_xl(name):
    maximum = get_max()
    minimum = get_min()
    avg = get_average()

    ws.title = "data"

    ws["C2"] = f"${avg}"
    ws["D2"] = f"${maximum}"
    ws["E2"] = f"${minimum}"
    wb.save(f"C:\\Users\\zacha\\OneDrive\\Documents\{name}.xlsx")
    print("file saved succesfully")

save_name = input("what would you like to name the spreadsheet you saved the data too? ")

#runs function
save_data_to_xl(save_name)
